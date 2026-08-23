from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from operandomerge.config import load_config, save_config
from operandomerge.export import export_csv_bundle, export_excel
from operandomerge.io import inspect_columns, read_table
from operandomerge.models import (
    AlignmentConfig,
    AlignmentMethod,
    ChannelConfig,
    DatasetConfig,
    DataType,
    DelayConfig,
    MergeConfig,
    TimeRepresentation,
)
from operandomerge.service import MergeService


def _full_config(path: Path) -> DatasetConfig:
    return DatasetConfig(
        path=path,
        name="instrument",
        time_column="clock",
        time_representation=TimeRepresentation.ELAPSED_MINUTES,
        channels=(ChannelConfig("signal", DataType.STEPWISE, "instrument_signal"),),
        alignment=AlignmentConfig(
            method=AlignmentMethod.REFERENCE_EVENT,
            manual_offset_s=2,
            source_event_time_s=10,
            target_event_time_s=20,
        ),
        delay=DelayConfig(manual_s=1, sampling_s=2, transport_s=3, dead_volume_s=4, analysis_s=5),
    )


def test_csv_and_xlsx_are_read_equivalently(table_writer) -> None:
    data = {"clock": [0, 1], "signal": [2.5, 3.5]}
    csv_path = table_writer("data_csv", data)
    xlsx_path = table_writer("data_excel", data, excel=True)
    assert inspect_columns(csv_path) == ["clock", "signal"]
    assert read_table(csv_path).equals(read_table(xlsx_path))


def test_config_roundtrip_preserves_every_gui_relevant_setting(table_writer, tmp_path) -> None:
    source = table_writer("roundtrip", {"clock": [0, 1], "signal": [2, 3]})
    original = _full_config(source)
    merge = MergeConfig(
        timeline="reference",
        reference_dataset="instrument",
        continuous_method="nearest",
        stepwise_method="none",
        exact_tolerance_s=0.01,
    )
    path = tmp_path / "analysis.json"
    save_config([original], merge, path)
    datasets, loaded_merge = load_config(path)
    assert datasets == [original]
    assert loaded_merge == merge


def test_excel_and_csv_exports_include_reproducibility_artifacts(table_writer, tmp_path) -> None:
    source = table_writer("export", {"clock": [0, 1], "signal": [2, 3]})
    config = DatasetConfig(
        path=source,
        name="instrument",
        time_column="clock",
        time_representation=TimeRepresentation.ELAPSED_SECONDS,
        channels=(ChannelConfig("signal"),),
    )
    result = MergeService().run([config])
    metadata = result.metadata.iloc[0]
    assert metadata["alignment_method"] == "elapsed"
    assert metadata["sampling_delay_s"] == 0
    assert result.configuration["datasets"][0]["channels"][0]["data_type"] == "continuous"
    excel = tmp_path / "report.xlsx"
    export_excel(result, excel)
    workbook = openpyxl.load_workbook(excel, read_only=True)
    assert workbook.sheetnames == ["merged", "metadata", "provenance", "qc", "configuration"]
    provenance_headers = [cell.value for cell in next(workbook["provenance"].iter_rows())]
    assert {"source_file", "source_column", "original_timestamp", "interpolation_method"} <= set(
        provenance_headers
    )
    bundle = tmp_path / "csv"
    export_csv_bundle(result, bundle)
    assert {path.name for path in bundle.iterdir()} == {
        "merged.csv",
        "metadata.csv",
        "provenance.csv",
        "qc.csv",
        "configuration.csv",
    }


def test_invalid_config_structure_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "bad.json"
    path.write_text(json.dumps({"merge": {}}), encoding="utf-8")
    try:
        load_config(path)
    except ValueError as error:
        assert "datasets" in str(error)
    else:
        raise AssertionError("invalid configuration was accepted")
