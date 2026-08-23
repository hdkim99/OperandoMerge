from __future__ import annotations

from dataclasses import replace
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from operandomerge.cli import main as cli_main
from operandomerge.config import load_config, save_config
from operandomerge.models import DelayConfig
from operandomerge.service import MergeService

REPOSITORY_ROOT = Path(__file__).parents[2]
SHOWCASE_DIRECTORY = REPOSITORY_ROOT / "examples" / "showcase"
SHOWCASE_CONFIG = SHOWCASE_DIRECTORY / "config.json"


def _showcase_result():
    datasets, merge_config = load_config(SHOWCASE_CONFIG)
    return datasets, merge_config, MergeService().run(datasets, merge_config)


def test_showcase_recovers_known_canonical_timeline_and_delays() -> None:
    _datasets, _merge, result = _showcase_result()
    np.testing.assert_allclose(result.merged["experiment_time_s"], np.arange(0.0, 601.0, 2.0))
    assert result.qc == []
    metadata = result.metadata.set_index("dataset")
    assert metadata.loc["reactor", "applied_offset_s"] == 0
    assert metadata.loc["mfc", "applied_offset_s"] == 20
    assert metadata.loc["ms", "transport_delay_s"] == 8
    assert metadata.loc["gc", "sampling_delay_s"] == 12
    assert metadata.loc["gc", "analysis_delay_s"] == 18
    assert metadata.loc["gc", "total_delay_s"] == 30
    assert metadata.loc["xrd", "applied_offset_s"] == -5
    assert metadata.loc["xrd", "analysis_delay_s"] == 15


def test_gc_samples_remain_discrete_and_physically_consistent() -> None:
    _datasets, _merge, result = _showcase_result()
    gc_columns = ["gc_co2_mol_pct", "gc_co_mol_pct", "gc_ch4_mol_pct"]
    measured = result.merged.dropna(subset=gc_columns, how="all")
    assert measured["experiment_time_s"].tolist() == [120, 300, 480]
    assert measured[gc_columns].sum(axis=1).tolist() == [100, 100, 100]
    assert result.merged["gc_co2_mol_pct"].isna().sum() == 298
    assert (
        result.merged.loc[
            result.merged["experiment_time_s"].isin([118, 122, 298, 302, 478, 482]),
            gc_columns,
        ]
        .isna()
        .all()
        .all()
    )
    provenance = result.provenance.query("output_column == 'gc_co2_mol_pct'")
    assert set(provenance["interpolation_method"]) == {"original"}
    assert provenance["total_delay_s"].tolist() == [30, 30, 30]
    assert provenance["original_timestamp"].tolist() == [
        "2026-01-01T12:02:30+00:00",
        "2026-01-01T12:05:30+00:00",
        "2026-01-01T12:08:30+00:00",
    ]


def test_showcase_boundary_semantics_do_not_extrapolate_or_invent_scans() -> None:
    _datasets, _merge, result = _showcase_result()
    timeline = result.merged.set_index("experiment_time_s")
    assert timeline.loc[0:18, "mfc_h2_flow_sccm"].isna().all()
    assert timeline.loc[20, "mfc_h2_flow_sccm"] == 20
    assert timeline.loc[120, "mfc_h2_flow_sccm"] == 30
    xrd_measured = timeline["xrd_lattice_parameter_A"].dropna()
    assert xrd_measured.index.tolist() == list(range(0, 601, 60))
    assert (
        timeline.loc[2, "xrd_reduced_phase_fraction"]
        == timeline.loc[0, "xrd_reduced_phase_fraction"]
    )


def test_interpolated_and_stepwise_provenance_reaches_real_source_rows() -> None:
    _datasets, _merge, result = _showcase_result()
    reactor = result.provenance.query(
        "output_column == 'reactor_temperature_C' and experiment_time_s == 2"
    ).iloc[0]
    assert reactor["interpolation_method"] == "linear"
    assert reactor["source_row_left"] == 0
    assert reactor["source_row_right"] == 1
    assert reactor["original_timestamp_left"] == "2026-01-01T12:00:00+00:00"
    assert reactor["original_timestamp_right"] == "2026-01-01T12:00:10+00:00"
    xrd = result.provenance.query(
        "output_column == 'xrd_reduced_phase_fraction' and experiment_time_s == 2"
    ).iloc[0]
    assert xrd["interpolation_method"] == "previous"
    assert xrd["source_row_left"] == xrd["source_row_right"] == 0


def test_showcase_config_roundtrip_preserves_alignment_semantics(tmp_path: Path) -> None:
    datasets, merge_config = load_config(SHOWCASE_CONFIG)
    saved = tmp_path / "roundtrip.json"
    save_config(datasets, merge_config, saved)
    loaded_datasets, loaded_merge = load_config(saved)
    assert loaded_datasets == datasets
    assert loaded_merge == merge_config
    assert loaded_datasets[1].alignment.effective_offset_s() == 20
    assert loaded_datasets[3].delay.total_s == 30
    assert loaded_datasets[4].alignment.manual_offset_s == -5


def test_physically_invalid_negative_showcase_delay_is_rejected() -> None:
    datasets, merge_config = load_config(SHOWCASE_CONFIG)
    datasets[2] = replace(datasets[2], delay=DelayConfig(transport_s=-8))
    with pytest.raises(ValueError, match="non-negative"):
        MergeService().run(datasets, merge_config)


def test_showcase_cli_writes_real_excel_and_alignment_plot(tmp_path: Path) -> None:
    report = tmp_path / "showcase.xlsx"
    figure = tmp_path / "alignment.png"
    assert (
        cli_main(
            [
                "merge",
                str(SHOWCASE_CONFIG),
                "--excel",
                str(report),
                "--plot",
                str(figure),
            ]
        )
        == 0
    )
    assert report.stat().st_size > 100_000
    assert figure.stat().st_size > 20_000
    workbook = openpyxl.load_workbook(report, read_only=True, data_only=True)
    assert workbook.sheetnames == ["merged", "metadata", "provenance", "qc", "configuration"]
    assert workbook["merged"].max_row == 302
    assert workbook["metadata"].max_row == 6
    merged_rows = list(workbook["merged"].iter_rows(values_only=True))
    header = list(merged_rows[0])
    time_index = header.index("experiment_time_s")
    gc_index = header.index("gc_co2_mol_pct")
    exported_gc = [
        (row[time_index], row[gc_index]) for row in merged_rows[1:] if row[gc_index] is not None
    ]
    assert exported_gc == [(120, 5), (300, 18), (480, 42)]
